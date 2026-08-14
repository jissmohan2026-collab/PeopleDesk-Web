


from dotenv import load_dotenv
load_dotenv()
from flask import Flask, render_template, request, redirect, url_for, flash, session
from flask_babel import Babel, _
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from config import Config
from models import db, bcrypt, User, Department, Consignment, PendingWork, Vehicle, SystemBalance, Road, FundManagement, LAC_ADF_Project, ApprovalSanctionTracking, ProjectExecution, PaymentUtilization, MonitoringInspection, CompletionAssetRegister, MLA_SDF_Document, MLA_SDF_Report
import uuid
from datetime import datetime
import os
import csv
import io
from flask import Response
import cloudinary
import cloudinary.uploader

app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)
bcrypt.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

def get_locale():
    return session.get('lang', request.accept_languages.best_match(app.config['LANGUAGES']))

babel = Babel(app, locale_selector=get_locale)

@app.route('/set_language/<lang>')
def set_language(lang):
    if lang in app.config['LANGUAGES']:
        session['lang'] = lang
    return redirect(request.referrer or url_for('index'))

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# --- Public Routes ---
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/track', methods=['GET', 'POST'])
def track():
    consignment = None
    if request.method == 'POST':
        tracking_id = request.form.get('tracking_id')
        consignment = Consignment.query.filter_by(tracking_id=tracking_id).first()
        if not consignment:
            flash(_('No consignment found with that tracking ID.), '), ')')
    return render_template('track.html', consignment=consignment)

@app.route('/contact')
def contact():
    return render_template('contact.html')

@app.route('/faq')
def faq():
    return render_template('faq.html')

@app.route('/privacy')
def privacy():
    return render_template('privacy.html')

@app.route('/terms')
def terms():
    return render_template('terms.html')

# --- Auth Routes ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        if current_user.role == 'superadmin':
            return redirect(url_for('superadmin_dashboard'))
        elif current_user.role == 'admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('user_dashboard'))

    if request.method == 'POST':
        try:
            username = request.form.get('username')
            password = request.form.get('password')
            user = User.query.filter_by(username=username).first()
            if user and user.check_password(password):
                login_user(user)
                if user.role == 'superadmin':
                    return redirect(url_for('superadmin_dashboard'))
                elif user.role == 'admin':
                    return redirect(url_for('admin_dashboard'))
                else:
                    return redirect(url_for('user_dashboard'))
            else:
                flash('Invalid credentials', 'danger')
        except Exception as e:
            flash(str(e), 'danger')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('user_dashboard'))
    if request.method == 'POST':
        try:
            username = request.form.get('username')
            password = request.form.get('password')
            name = request.form.get('name')
            address = request.form.get('address')
            phone_no = request.form.get('phone_no')
            place = request.form.get('place')
            district = request.form.get('district')
            
            if User.query.filter_by(username=username).first():
                flash(_('Username already exists.), '), ')')
                return redirect(url_for('register'))
            if phone_no and User.query.filter_by(phone_no=phone_no).first():
                flash(_('Phone number already registered.), '), ')')
                return redirect(url_for('register'))
                
            new_user = User(
                username=username, role='user',
                name=name, address=address, phone_no=phone_no, place=place, district=district
            )
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.commit()
            login_user(new_user)
            return redirect(url_for('user_dashboard'))
        except Exception as e:
            flash(str(e), 'danger')
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

# --- User Routes ---
@app.route('/user/dashboard')
@login_required
def user_dashboard():
    if current_user.role != 'user':
        return redirect(url_for('index'))
    consignments = Consignment.query.filter_by(user_id=current_user.id).all()
    return render_template('user/dashboard.html', consignments=consignments)

@app.route('/user/submit', methods=['GET', 'POST'])
@login_required
def submit_issue():
    if current_user.role != 'user':
        return redirect(url_for('index'))
    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        department_id = request.form.get('department_id')
        
        attachment_url = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                try:
                    upload_result = cloudinary.uploader.upload(file)
                    attachment_url = upload_result.get('secure_url')
                except Exception as e:
                    flash(f'File upload failed: {str(e)}', 'danger')
        
        tracking_id = "CCN" + str(uuid.uuid4().int)[:9] # Generate a simple tracking ID
        new_con = Consignment(title=title, description=description, tracking_id=tracking_id, 
                              user_id=current_user.id, department_id=department_id, attachment_url=attachment_url)
        db.session.add(new_con)
        db.session.commit()
        flash(f'Application submitted successfully! Your tracking ID is {tracking_id}', 'success')
        return redirect(url_for('user_dashboard'))
    
    departments = Department.query.all()
    return render_template('user/submit.html', departments=departments)

# --- Admin Routes ---
@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    status_filter = request.args.get('status')
    date_filter = request.args.get('date')
    
    query = Consignment.query.filter_by(assigned_admin_id=current_user.id)
    
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    consignments = query.order_by(Consignment.created_at.desc()).all()
    
    # Python-level date filtering if date_filter is provided (format: YYYY-MM-DD)
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            consignments = [con for con in consignments if con.created_at and con.created_at.date() == filter_date]
        except ValueError:
            pass

    return render_template('admin/dashboard.html', consignments=consignments, status_filter=status_filter, date_filter=date_filter)

@app.route('/admin/profile', methods=['GET', 'POST'])
@login_required
def admin_profile():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    if request.method == 'POST':
        new_password = request.form.get('password')
        if new_password:
            current_user.set_password(new_password)
            db.session.commit()
            flash(_('Password updated successfully.), '), ')')
            return redirect(url_for('admin_profile'))
    return render_template('admin/profile.html')

@app.route('/admin/contact')
@login_required
def admin_contact():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    return render_template('admin/contact.html')

@app.route('/admin/create_application', methods=['GET', 'POST'])
@login_required
def admin_create_application():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        department_id = request.form.get('department_id')
        
        phone_no = request.form.get('phone_no')
        name = request.form.get('name')
        address = request.form.get('address')
        place = request.form.get('place')
        district = request.form.get('district')
        
        user = None
        if phone_no:
            user = User.query.filter_by(phone_no=phone_no).first()
            if not user:
                username = phone_no
                user = User(username=username, role='user', name=name, address=address, phone_no=phone_no, place=place, district=district)
                user.set_password(phone_no)
                db.session.add(user)
                db.session.flush()
                flash(_('New user account created automatically.), '), ')')
        
        user_id = user.id if user else current_user.id
        
        attachment_url = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                try:
                    upload_result = cloudinary.uploader.upload(file)
                    attachment_url = upload_result.get('secure_url')
                except Exception as e:
                    flash(f'File upload failed: {str(e)}', 'danger')
        
        tracking_id = "CCN" + str(uuid.uuid4().int)[:9]
        new_con = Consignment(title=title, description=description, tracking_id=tracking_id, 
                              user_id=user_id, department_id=department_id, attachment_url=attachment_url)
        db.session.add(new_con)
        db.session.commit()
        flash(f'Application submitted successfully! Tracking ID: {tracking_id}', 'success')
        return redirect(url_for('admin_dashboard'))
    
    departments = Department.query.all()
    return render_template('admin/create_application.html', departments=departments)

@app.route('/admin/update_status/<int:id>', methods=['POST'])
@login_required
def admin_update_status(id):
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    consignment = Consignment.query.get_or_404(id)
    if consignment.assigned_admin_id != current_user.id:
        flash(_('Not authorized.), '), ')')
        return redirect(url_for('admin_dashboard'))
    
    new_status = request.form.get('status')
    if new_status in ['Pending', 'Under Review', 'Approved', 'Completed', 'Closed']:
        consignment.status = new_status
        db.session.commit()
        flash(_('Status updated.), '), ')')
    return redirect(url_for('admin_dashboard'))

# --- SuperAdmin Routes ---
@app.route('/superadmin/dashboard')
@login_required
def superadmin_dashboard():
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
    
    # Ensure a SystemBalance row exists
    balance = SystemBalance.query.order_by(SystemBalance.id.asc()).first()
    if not balance:
        balance = SystemBalance(total_balance=10000000.0, current_balance=10000000.0, used_up_balance=0.0)
        db.session.add(balance)
        db.session.commit()
    
    # Calculate values from new FundManagement table
    allocated = db.session.query(db.func.sum(FundManagement.annual_fund_allocation)).scalar() or 0.0
    sanctioned = db.session.query(db.func.sum(FundManagement.amount_sanctioned)).scalar() or 0.0
    released = db.session.query(db.func.sum(FundManagement.amount_released)).scalar() or 0.0
    utilized = db.session.query(db.func.sum(FundManagement.amount_utilized)).scalar() or 0.0
    balance_left = db.session.query(db.func.sum(FundManagement.balance_amount)).scalar() or 0.0
    
    # If no records exist in FundManagement, fallback to SystemBalance details or zero
    if allocated == 0.0:
        allocated = db.session.query(db.func.sum(LAC_ADF_Project.estimated_project_cost)).scalar() or 0.0
        sanctioned = db.session.query(db.func.sum(LAC_ADF_Project.mla_adf_amount)).scalar() or 0.0
        released = sanctioned * 0.8
        utilized = db.session.query(db.func.sum(ProjectExecution.amount_paid)).scalar() or 0.0
        balance_left = balance.total_balance - utilized

    # Calculate projects counts from LAC_ADF_Project
    total_projects = LAC_ADF_Project.query.count()
    completed_projects = LAC_ADF_Project.query.filter_by(project_status='Completed').count()
    ongoing_projects = LAC_ADF_Project.query.filter(LAC_ADF_Project.project_status.in_(['Approved', 'Execution'])).count()
    pending_projects = LAC_ADF_Project.query.filter(LAC_ADF_Project.project_status.in_(['Proposed', 'Delayed'])).count()
    
    if total_projects == 0:
        # Fallback to legacy count if new tables are empty on setup
        total_projects = PendingWork.query.count() + Road.query.count()
        completed_projects = PendingWork.query.filter(PendingWork.status.in_(['Complete', 'Completed'])).count() + Road.query.filter(Road.status.in_(['Complete', 'Completed'])).count()
        ongoing_projects = PendingWork.query.filter(PendingWork.status.in_(['Pending', 'In Progress'])).count() + Road.query.filter(Road.status.in_(['Pending', 'In Progress'])).count()
        pending_projects = total_projects - completed_projects

    # Calculate average progress
    avg_physical = db.session.query(db.func.avg(ProjectExecution.physical_progress_pct)).scalar() or 0.0
    avg_financial = db.session.query(db.func.avg(ProjectExecution.financial_progress_pct)).scalar() or 0.0
    
    if avg_physical == 0.0 and total_projects > 0:
        avg_physical = (completed_projects / total_projects) * 100
    if avg_financial == 0.0 and allocated > 0:
        avg_financial = (utilized / allocated) * 100

    # Update System Balance dynamically
    balance.used_up_balance = utilized
    balance.current_balance = balance.total_balance - utilized
    db.session.commit()

    stats = {
        'total_apps': Consignment.query.count(),
        'pending_apps': Consignment.query.filter_by(status='Pending').count(),
        'approved_apps': Consignment.query.filter_by(status='Approved').count(),
        'completed_apps': Consignment.query.filter_by(status='Completed').count(),
        'total_users': User.query.filter_by(role='user').count(),
        'total_admins': User.query.filter(User.role.in_(['admin', 'superadmin'])).count(),
        
        # New MLA-SDF / LAC-ADF metrics
        'allocated': allocated,
        'sanctioned': sanctioned,
        'released': released,
        'utilized': utilized,
        'balance_left': balance_left,
        'total_projects': total_projects,
        'completed_projects': completed_projects,
        'ongoing_projects': ongoing_projects,
        'pending_projects': pending_projects,
        'physical_progress': avg_physical,
        'financial_progress': avg_financial
    }
    
    return render_template('superadmin/dashboard.html', stats=stats, balance=balance)

@app.route('/superadmin/update_balance', methods=['POST'])
@login_required
def superadmin_update_balance():
    if current_user.role != 'superadmin':
        return {"status": "error", "message": "Unauthorized"}, 403
        
    balance = SystemBalance.query.order_by(SystemBalance.id.asc()).first()
    if not balance:
        balance = SystemBalance()
        db.session.add(balance)
        
    try:
        data = request.get_json()
        if not data:
            return {"status": "error", "message": "Invalid data"}, 400
            
        if 'total_balance' in data:
            balance.total_balance = float(data['total_balance'])
        if 'current_balance' in data:
            balance.current_balance = float(data['current_balance'])
            
        balance.used_up_balance = balance.total_balance - balance.current_balance
            
        db.session.commit()
        flash(_('Balance updated successfully.'))
        return {"status": "success", "message": "Balance updated successfully"}
    except Exception as e:
        db.session.rollback()
        return {"status": "error", "message": str(e)}, 500

@app.route('/superadmin/applications')
@login_required
def superadmin_applications():
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
    
    status_filter = request.args.get('status')
    date_filter = request.args.get('date')
    search_query = request.args.get('search', '').strip()
    
    query = Consignment.query
    
    if status_filter:
        query = query.filter_by(status=status_filter)
        
    consignments = query.order_by(Consignment.created_at.desc()).all()
    
    if search_query:
        filtered_consignments = []
        for con in consignments:
            match = False
            if search_query.lower() in con.tracking_id.lower() or (con.title and search_query.lower() in con.title.lower()):
                match = True
            elif con.submitter:
                if con.submitter.phone_no and search_query in con.submitter.phone_no:
                    match = True
                elif con.submitter.username and search_query.lower() in con.submitter.username.lower():
                    match = True
                elif con.submitter.name and search_query.lower() in con.submitter.name.lower():
                    match = True
            if match:
                filtered_consignments.append(con)
        consignments = filtered_consignments

    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            consignments = [con for con in consignments if con.created_at and con.created_at.date() == filter_date]
        except ValueError:
            pass
            
    admins = User.query.filter_by(role='admin').all()
    
    return render_template('superadmin/applications.html', consignments=consignments, admins=admins, status_filter=status_filter, date_filter=date_filter, search_query=search_query)

@app.route('/superadmin/users', methods=['GET', 'POST'])
@login_required
def superadmin_users():
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
    
    search_query = request.args.get('search', '').strip()
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role', 'user')
        department_id = request.form.get('department_id') or None
        
        if User.query.filter_by(username=username).first():
            flash(_('Username already exists.'), 'danger')
        else:
            new_user = User(username=username, role=role, department_id=department_id)
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.commit()
            flash(_('User created successfully.'), 'success')
        return redirect(url_for('superadmin_users'))
        
    query = User.query
    if search_query:
        query = query.filter(db.or_(
            User.username.ilike(f'%{search_query}%'),
            User.name.ilike(f'%{search_query}%'),
            User.phone_no.ilike(f'%{search_query}%')
        ))
    users = query.all()
    departments = Department.query.all()
    return render_template('superadmin/users.html', users=users, departments=departments, search_query=search_query)

@app.route('/superadmin/edit_user/<int:user_id>', methods=['POST'])
@login_required
def superadmin_edit_user(user_id):
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
        
    user = User.query.get_or_404(user_id)
    new_password = request.form.get('password')
    department_id = request.form.get('department_id') or None
    
    if new_password:
        user.set_password(new_password)
    user.department_id = department_id
    db.session.commit()
    flash(f'User {user.username} updated successfully.', 'success')
    return redirect(url_for('superadmin_users'))

@app.route('/superadmin/departments', methods=['GET', 'POST'])
@login_required
def superadmin_departments():
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        name = request.form.get('name')
        officer_name = request.form.get('officer_name')
        officer_phone = request.form.get('officer_phone')
        officer_email = request.form.get('officer_email')
        if Department.query.filter_by(name=name).first():
            flash(_('Department already exists.'), 'danger')
        else:
            new_dept = Department(name=name, officer_name=officer_name, officer_phone=officer_phone, officer_email=officer_email)
            db.session.add(new_dept)
            db.session.commit()
            flash(_('Department created successfully.'), 'success')
        return redirect(url_for('superadmin_departments'))
        
    departments = Department.query.all()
    return render_template('superadmin/departments.html', departments=departments)

@app.route('/superadmin/delete_department/<int:dept_id>', methods=['POST'])
@login_required
def superadmin_delete_department(dept_id):
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
    dept = Department.query.get_or_404(dept_id)
    User.query.filter_by(department_id=dept.id).update({User.department_id: None})
    Consignment.query.filter_by(department_id=dept.id).update({Consignment.department_id: None})
    db.session.delete(dept)
    db.session.commit()
    flash(_('Department deleted successfully.'), 'success')
    return redirect(url_for('superadmin_departments'))

@app.route('/superadmin/user/<int:user_id>')
@login_required
def superadmin_user_profile(user_id):
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
    user = User.query.get_or_404(user_id)
    return render_template('superadmin/user_profile.html', user=user)

@app.route('/superadmin/delete_user/<int:user_id>', methods=['POST'])
@login_required
def superadmin_delete_user(user_id):
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
    if current_user.id == user_id:
        flash(_('You cannot delete your own profile.'), 'danger')
        return redirect(url_for('superadmin_users'))
    user = User.query.get_or_404(user_id)
    Consignment.query.filter_by(user_id=user.id).update({Consignment.user_id: None})
    Consignment.query.filter_by(assigned_admin_id=user.id).update({Consignment.assigned_admin_id: None})
    db.session.delete(user)
    db.session.commit()
    flash(_('User deleted successfully.'), 'success')
    return redirect(url_for('superadmin_users'))

@app.route('/superadmin/delete_application/<int:con_id>', methods=['POST'])
@login_required
def superadmin_delete_application(con_id):
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
    consignment = Consignment.query.get_or_404(con_id)
    db.session.delete(consignment)
    db.session.commit()
    flash(_('Application deleted successfully.'), 'success')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/superadmin/edit_application/<int:con_id>', methods=['GET', 'POST'])
@login_required
def superadmin_edit_application(con_id):
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
    consignment = Consignment.query.get_or_404(con_id)
    if request.method == 'POST':
        consignment.title = request.form.get('title')
        consignment.description = request.form.get('description')
        db.session.commit()
        flash(_('Application updated successfully.'), 'success')
        return redirect(url_for('superadmin_dashboard'))
    return render_template('superadmin/edit_application.html', consignment=consignment)

@app.route('/superadmin/view_application/<int:con_id>')
@login_required
def superadmin_view_application(con_id):
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
    consignment = Consignment.query.get_or_404(con_id)
    return render_template('superadmin/view_application.html', consignment=consignment)

@app.route('/superadmin/export/csv')
@login_required
def superadmin_export_csv():
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
    consignments = Consignment.query.all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Tracking ID', 'Title', 'Description', 'Status', 'User Name', 'User Phone', 'Department', 'Created At'])
    for c in consignments:
        user_name = c.submitter.name if c.submitter else 'N/A'
        user_phone = c.submitter.phone_no if c.submitter else 'N/A'
        dept_name = c.department.name if c.department else 'N/A'
        writer.writerow([c.id, c.tracking_id, c.title, c.description, c.status, user_name, user_phone, dept_name, c.created_at])
        
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-disposition": "attachment; filename=applications.csv"}
    )

@app.route('/superadmin/export/xml')
@login_required
def superadmin_export_xml():
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
    consignments = Consignment.query.all()
    
    xml_data = '<?xml version="1.0" encoding="UTF-8"?>\n<Applications>\n'
    for c in consignments:
        user_name = c.submitter.name if c.submitter else 'N/A'
        user_phone = c.submitter.phone_no if c.submitter else 'N/A'
        dept_name = c.department.name if c.department else 'N/A'
        xml_data += f'''  <Application>
    <ID>{c.id}</ID>
    <TrackingID>{c.tracking_id}</TrackingID>
    <Title>{c.title}</Title>
    <Status>{c.status}</Status>
    <UserName>{user_name}</UserName>
    <UserPhone>{user_phone}</UserPhone>
    <Department>{dept_name}</Department>
  </Application>\n'''
    xml_data += '</Applications>'
    
    return Response(
        xml_data,
        mimetype="application/xml",
        headers={"Content-disposition": "attachment; filename=applications.xml"}
    )

@app.route('/superadmin/assign/<int:con_id>', methods=['POST'])
@login_required
def superadmin_assign(con_id):
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
    consignment = Consignment.query.get_or_404(con_id)
    admin_id = request.form.get('admin_id')
    if admin_id:
        consignment.assigned_admin_id = admin_id
        db.session.commit()
        flash(_('Admin assigned successfully.'))
    return redirect(url_for('superadmin_dashboard'))

from flask import send_file
from openpyxl import Workbook
from io import BytesIO

@app.route('/pending-works', methods=['GET'])
@login_required
def pending_works():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    
    status_filter = request.args.get('status', '')
    local_body_filter = request.args.get('local_body', '')
    ward_filter = request.args.get('ward_no', '')
    amount_min = request.args.get('amount_min', '')
    amount_max = request.args.get('amount_max', '')
    
    query = PendingWork.query
    if status_filter:
        query = query.filter(PendingWork.status == status_filter)
    if local_body_filter:
        query = query.filter(PendingWork.local_body.ilike(f"%{local_body_filter}%"))
    if ward_filter:
        try:
            query = query.filter(PendingWork.ward_no == int(ward_filter))
        except ValueError:
            pass
    if amount_min:
        try:
            query = query.filter(PendingWork.amount >= int(amount_min))
        except ValueError:
            pass
    if amount_max:
        try:
            query = query.filter(PendingWork.amount <= int(amount_max))
        except ValueError:
            pass
            
    works = query.all()
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    return render_template('admin/pending_works.html', works=works, 
                           status_filter=status_filter, local_body_filter=local_body_filter,
                           ward_filter=ward_filter, amount_min=amount_min, amount_max=amount_max,
                           base_template=base_template)

@app.route('/pending-works/add', methods=['POST'])
@login_required
def add_pending_work():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
        
    try:
        work_name = request.form.get('work_name')
        amount = int(request.form.get('amount', 0))
        installment_1 = int(request.form.get('installment_1', 0))
        installment_2 = int(request.form.get('installment_2', 0))
        installment_3 = int(request.form.get('installment_3', 0))
        balance_amount = amount - installment_1 - installment_2 - installment_3
        local_body = request.form.get('local_body')
        ward_no = int(request.form.get('ward_no', 0))
        status = request.form.get('status', 'Pending')
        file_status = request.form.get('file_status', '')
        department = request.form.get('department', '')
        remarks = request.form.get('remarks', '')
        contractor_name = request.form.get('contractor_name', '')
        contractor_address = request.form.get('contractor_address', '')
        contractor_phone = request.form.get('contractor_phone', '')
        
        work = PendingWork(
            work_name=work_name,
            amount=amount,
            installment_1=installment_1,
            installment_2=installment_2,
            installment_3=installment_3,
            balance_amount=balance_amount,
            local_body=local_body,
            ward_no=ward_no,
            status=status,
            file_status=file_status,
            department=department,
            remarks=remarks,
            contractor_name=contractor_name,
            contractor_address=contractor_address,
            contractor_phone=contractor_phone
        )
        db.session.add(work)
        db.session.commit()
        flash(_('Project Work added successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding work: {str(e)}', 'danger')
        
    return redirect(url_for('pending_works'))

@app.route('/pending-works/edit/<int:work_id>', methods=['POST'])
@login_required
def edit_pending_work(work_id):
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
        
    work = PendingWork.query.get_or_404(work_id)
    try:
        work.work_name = request.form.get('work_name')
        work.amount = int(request.form.get('amount', 0))
        work.installment_1 = int(request.form.get('installment_1', 0))
        work.installment_2 = int(request.form.get('installment_2', 0))
        work.installment_3 = int(request.form.get('installment_3', 0))
        work.balance_amount = work.amount - work.installment_1 - work.installment_2 - work.installment_3
        work.local_body = request.form.get('local_body')
        work.ward_no = int(request.form.get('ward_no', 0))
        work.status = request.form.get('status')
        work.file_status = request.form.get('file_status', '')
        work.department = request.form.get('department', '')
        work.remarks = request.form.get('remarks', '')
        work.contractor_name = request.form.get('contractor_name', '')
        work.contractor_address = request.form.get('contractor_address', '')
        work.contractor_phone = request.form.get('contractor_phone', '')
        
        db.session.commit()
        flash(_('Project Work updated successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating work: {str(e)}', 'danger')
        
    return redirect(url_for('pending_works'))

@app.route('/pending-works/download', methods=['GET'])
@login_required
def download_pending_works():
    if current_user.role != 'superadmin':
        flash(_('Unauthorized access.'), 'danger')
        return redirect(url_for('index'))
        
    status_filter = request.args.get('status', '')
    local_body_filter = request.args.get('local_body', '')
    ward_filter = request.args.get('ward_no', '')
    amount_min = request.args.get('amount_min', '')
    amount_max = request.args.get('amount_max', '')
    
    query = PendingWork.query
    if status_filter:
        query = query.filter(PendingWork.status == status_filter)
    if local_body_filter:
        query = query.filter(PendingWork.local_body.ilike(f"%{local_body_filter}%"))
    if ward_filter:
        try:
            query = query.filter(PendingWork.ward_no == int(ward_filter))
        except ValueError:
            pass
    if amount_min:
        try:
            query = query.filter(PendingWork.amount >= int(amount_min))
        except ValueError:
            pass
    if amount_max:
        try:
            query = query.filter(PendingWork.amount <= int(amount_max))
        except ValueError:
            pass
            
    works = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Works"
    
    headers = ["SR (ID)", "Work Name", "Amount", "Installment 1", "Installment 2", "Installment 3", "Balance Amount", "Local Body", "Ward No", "Status", "File Status", "Department", "Contractor Name", "Contractor Address", "Contractor Phone", "Remarks"]
    ws.append(headers)
    
    for idx, w in enumerate(works, start=1):
        ws.append([idx, w.work_name, w.amount, w.installment_1 or 0, w.installment_2 or 0, w.installment_3 or 0, w.balance_amount or 0, w.local_body, w.ward_no, w.status, w.file_status, w.department, w.contractor_name or '', w.contractor_address or '', w.contractor_phone or '', w.remarks])
        
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    filename = f"pending_works_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = Response(
        file_stream.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route('/roads', methods=['GET'])
@login_required
def roads():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    
    status_filter = request.args.get('status', '')
    constituency_filter = request.args.get('constituency', '')
    road_name_filter = request.args.get('road_name', '')
    
    query = Road.query
    if status_filter:
        query = query.filter(Road.status == status_filter)
    if constituency_filter:
        query = query.filter(Road.constituency.ilike(f'%{constituency_filter}%'))
    if road_name_filter:
        query = query.filter(Road.road_name.ilike(f'%{road_name_filter}%'))
        
    roads_list = query.order_by(Road.id.desc()).all()
    
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    return render_template('admin/roads.html', roads=roads_list, 
                           base_template=base_template,
                           status_filter=status_filter,
                           constituency_filter=constituency_filter,
                           road_name_filter=road_name_filter)

@app.route('/roads/add', methods=['POST'])
@login_required
def add_road():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
        
    try:
        constituency = request.form.get('constituency')
        mla_name = request.form.get('mla_name')
        local_govt_name = request.form.get('local_govt_name')
        local_govt_type = request.form.get('local_govt_type')
        road_name = request.form.get('road_name')
        road_width = float(request.form.get('road_width', 0.0))
        road_length = float(request.form.get('road_length', 0.0))
        estimate_cost = float(request.form.get('estimate_cost', 0.0))
        status = request.form.get('status', 'Pending')
        remarks = request.form.get('remarks', '')
        approval_date = request.form.get('approval_date', '')
        installment_1 = int(request.form.get('installment_1', 0))
        installment_2 = int(request.form.get('installment_2', 0))
        installment_3 = int(request.form.get('installment_3', 0))
        contractor_name = request.form.get('contractor_name', '')
        contractor_address = request.form.get('contractor_address', '')
        contractor_phone = request.form.get('contractor_phone', '')

        attachment_url = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                try:
                    upload_result = cloudinary.uploader.upload(file)
                    attachment_url = upload_result.get('secure_url')
                except Exception as e:
                    flash(f'File upload failed: {str(e)}', 'danger')
        
        road = Road(
            constituency=constituency,
            mla_name=mla_name,
            local_govt_name=local_govt_name,
            local_govt_type=local_govt_type,
            road_name=road_name,
            road_width=road_width,
            road_length=road_length,
            estimate_cost=estimate_cost,
            status=status,
            remarks=remarks,
            approval_date=approval_date,
            installment_1=installment_1,
            installment_2=installment_2,
            installment_3=installment_3,
            contractor_name=contractor_name,
            contractor_address=contractor_address,
            contractor_phone=contractor_phone,
            attachment_url=attachment_url
        )
        db.session.add(road)
        db.session.commit()
        flash(_('Project added successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding project: {str(e)}', 'danger')
        
    return redirect(url_for('roads'))

@app.route('/roads/edit/<int:road_id>', methods=['POST'])
@login_required
def edit_road(road_id):
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
        
    road = Road.query.get_or_404(road_id)
    try:
        road.constituency = request.form.get('constituency')
        road.mla_name = request.form.get('mla_name')
        road.local_govt_name = request.form.get('local_govt_name')
        road.local_govt_type = request.form.get('local_govt_type')
        road.road_name = request.form.get('road_name')
        road.road_width = float(request.form.get('road_width', 0.0))
        road.road_length = float(request.form.get('road_length', 0.0))
        road.estimate_cost = float(request.form.get('estimate_cost', 0.0))
        road.status = request.form.get('status')
        road.remarks = request.form.get('remarks', '')
        road.approval_date = request.form.get('approval_date', '')
        road.installment_1 = int(request.form.get('installment_1', 0))
        road.installment_2 = int(request.form.get('installment_2', 0))
        road.installment_3 = int(request.form.get('installment_3', 0))
        road.contractor_name = request.form.get('contractor_name', '')
        road.contractor_address = request.form.get('contractor_address', '')
        road.contractor_phone = request.form.get('contractor_phone', '')

        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                try:
                    upload_result = cloudinary.uploader.upload(file)
                    road.attachment_url = upload_result.get('secure_url')
                except Exception as e:
                    flash(f'File upload failed: {str(e)}', 'danger')
        
        db.session.commit()
        flash(_('Project updated successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating project: {str(e)}', 'danger')
        
    return redirect(url_for('roads'))


@app.route('/roads/delete_attachment/<int:road_id>', methods=['POST'])
@login_required
def delete_road_attachment(road_id):
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
        
    road = Road.query.get_or_404(road_id)
    try:
        road.attachment_url = None
        db.session.commit()
        flash(_('Attachment deleted successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting attachment: {str(e)}', 'danger')
        
    return redirect(url_for('roads'))

@app.route('/roads/download', methods=['GET'])
@login_required
def download_roads():
    if current_user.role != 'superadmin':
        return redirect(url_for('index'))
        
    status_filter = request.args.get('status', '')
    constituency_filter = request.args.get('constituency', '')
    road_name_filter = request.args.get('road_name', '')
    
    query = Road.query
    if status_filter:
        query = query.filter(Road.status == status_filter)
    if constituency_filter:
        query = query.filter(Road.constituency.ilike(f'%{constituency_filter}%'))
    if road_name_filter:
        query = query.filter(Road.road_name.ilike(f'%{road_name_filter}%'))
        
    roads_list = query.order_by(Road.id.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Roads"
    
    headers = ["SR (ID)", "Road Name", "Constituency", "MLA Name", "Local Self Govt", "Nature of Self Govt", "Width (m)", "Length (m)", "Estimate Cost (₹)", "Status", "Approval Date", "Installment 1", "Installment 2", "Installment 3", "Contractor Name", "Contractor Address", "Contractor Phone", "Attachment URL", "Remarks"]
    ws.append(headers)
    
    for idx, r in enumerate(roads_list, start=1):
        ws.append([idx, r.road_name, r.constituency, r.mla_name, r.local_govt_name, r.local_govt_type, r.road_width, r.road_length, r.estimate_cost, r.status, r.approval_date or '', r.installment_1 or 0, r.installment_2 or 0, r.installment_3 or 0, r.contractor_name or '', r.contractor_address or '', r.contractor_phone or '', r.attachment_url or '', r.remarks])
        
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    filename = f"roads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = Response(
        file_stream.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route('/vehicles', methods=['GET'])
@login_required
def vehicles():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
        
    status_filter = request.args.get('status', '')
    type_filter = request.args.get('vehicle_type', '')
    
    query = Vehicle.query
    if status_filter:
        query = query.filter(Vehicle.status == status_filter)
    if type_filter:
        query = query.filter(Vehicle.vehicle_type == type_filter)
        
    vehicles_list = query.all()
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    return render_template('admin/vehicles.html', vehicles=vehicles_list, 
                           status_filter=status_filter, type_filter=type_filter,
                           base_template=base_template)

@app.route('/vehicles/add', methods=['POST'])
@login_required
def add_vehicle():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
        
    try:
        vehicle_number = request.form.get('vehicle_number')
        vehicle_model = request.form.get('vehicle_model')
        vehicle_year = int(request.form.get('vehicle_year', 0))
        driver_name = request.form.get('driver_name')
        driver_contract = request.form.get('driver_contract')
        driver_license = request.form.get('driver_license')
        vehicle_type = request.form.get('vehicle_type', 'owned')
        status = request.form.get('status', 'Pending')
        
        veh = Vehicle(
            vehicle_number=vehicle_number,
            vehicle_model=vehicle_model,
            vehicle_year=vehicle_year,
            driver_name=driver_name,
            driver_contract=driver_contract,
            driver_license=driver_license,
            vehicle_type=vehicle_type,
            status=status
        )
        db.session.add(veh)
        db.session.commit()
        flash(_('Vehicle added successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding vehicle: {str(e)}', 'danger')
        
    return redirect(url_for('vehicles'))

@app.route('/vehicles/edit/<int:veh_id>', methods=['POST'])
@login_required
def edit_vehicle(veh_id):
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
        
    veh = Vehicle.query.get_or_404(veh_id)
    try:
        veh.vehicle_number = request.form.get('vehicle_number')
        veh.vehicle_model = request.form.get('vehicle_model')
        veh.vehicle_year = int(request.form.get('vehicle_year', 0))
        veh.driver_name = request.form.get('driver_name')
        veh.driver_contract = request.form.get('driver_contract')
        veh.driver_license = request.form.get('driver_license')
        veh.vehicle_type = request.form.get('vehicle_type')
        veh.status = request.form.get('status')
        
        db.session.commit()
        flash(_('Vehicle updated successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating vehicle: {str(e)}', 'danger')
        
    return redirect(url_for('vehicles'))


@app.route('/mla-sdf/fund-management')
@login_required
def mla_sdf_fund_management():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    funds = FundManagement.query.all()
    if not funds:
        default_fund = FundManagement(
            financial_year="2026-27",
            constituency="Trivandrum",
            mla_name="MLA Name",
            fund_type="LAC-ADF",
            annual_fund_allocation=5000000.0,
            previous_year_balance=1200000.0,
            total_available_fund=6200000.0,
            amount_recommended=4500000.0,
            amount_sanctioned=4000000.0,
            amount_released=3500000.0,
            amount_utilized=3000000.0,
            balance_amount=3200000.0,
            number_of_projects=8,
            pending_projects=5,
            completed_projects=3
        )
        db.session.add(default_fund)
        db.session.commit()
        funds = [default_fund]
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    return render_template('admin/mla_sdf/fund_management.html', base_template=base_template, funds=funds)


@app.route('/mla-sdf/fund-management/add', methods=['POST'])
@login_required
def add_fund_management():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    try:
        financial_year = request.form.get('financial_year')
        constituency = request.form.get('constituency')
        mla_name = request.form.get('mla_name')
        fund_type = request.form.get('fund_type')
        annual_fund_allocation = float(request.form.get('annual_fund_allocation', 0.0))
        previous_year_balance = float(request.form.get('previous_year_balance', 0.0))
        amount_recommended = float(request.form.get('amount_recommended', 0.0))
        amount_sanctioned = float(request.form.get('amount_sanctioned', 0.0))
        amount_released = float(request.form.get('amount_released', 0.0))
        amount_utilized = float(request.form.get('amount_utilized', 0.0))
        number_of_projects = int(request.form.get('number_of_projects', 0))
        pending_projects = int(request.form.get('pending_projects', 0))
        completed_projects = int(request.form.get('completed_projects', 0))
        
        total_available_fund = annual_fund_allocation + previous_year_balance
        balance_amount = total_available_fund - amount_utilized
        
        fund = FundManagement(
            financial_year=financial_year,
            constituency=constituency,
            mla_name=mla_name,
            fund_type=fund_type,
            annual_fund_allocation=annual_fund_allocation,
            previous_year_balance=previous_year_balance,
            total_available_fund=total_available_fund,
            amount_recommended=amount_recommended,
            amount_sanctioned=amount_sanctioned,
            amount_released=amount_released,
            amount_utilized=amount_utilized,
            balance_amount=balance_amount,
            number_of_projects=number_of_projects,
            pending_projects=pending_projects,
            completed_projects=completed_projects
        )
        db.session.add(fund)
        db.session.commit()
        flash(_('Fund record added successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding fund record: {str(e)}', 'danger')
        
    return redirect(url_for('mla_sdf_fund_management'))

@app.route('/mla-sdf/proposals')
@login_required
def mla_sdf_proposals():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    projects = LAC_ADF_Project.query.all()
    if not projects:
        default_project = LAC_ADF_Project(
            project_name="Constructing Community Hall Block",
            project_category="Building",
            project_description="Durable asset for public meetings",
            constituency="Trivandrum",
            local_body="Corporation",
            ward="12",
            location="East Fort",
            beneficiary_institution="Public Library",
            beneficiary_details="Local community",
            mla_recommendation_date="2026-04-12",
            estimated_project_cost=2500000.0,
            mla_adf_amount=2000000.0,
            other_fund_contribution=500000.0,
            implementing_department="PWD",
            implementing_agency="Contractor A",
            project_officer="Officer John",
            priority="High",
            project_status="Proposed"
        )
        db.session.add(default_project)
        db.session.commit()
        projects = [default_project]
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    return render_template('admin/mla_sdf/proposals.html', base_template=base_template, projects=projects)

@app.route('/mla-sdf/approvals')
@login_required
def mla_sdf_approvals():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    trackings = ApprovalSanctionTracking.query.all()
    if not trackings:
        project = LAC_ADF_Project.query.first()
        if project:
            default_tracking = ApprovalSanctionTracking(
                project_id=project.id,
                administrative_department="Finance",
                mla_recommendation="Recommended",
                administrative_sanction_no="AS-102-2026",
                administrative_sanction_date="2026-05-10",
                technical_sanction_no="TS-554-2026",
                technical_sanction_date="2026-05-15",
                financial_concurrence_no="FC-77-2026",
                financial_concurrence_date="2026-05-20",
                detailed_estimate_prepared=True,
                tender_quotation_required=True,
                tender_date="2026-06-01",
                work_order_no="WO-9988",
                work_order_date="2026-06-10",
                agreement_no="AG-2026-9",
                agreement_date="2026-06-15"
            )
            db.session.add(default_tracking)
            db.session.commit()
            trackings = [default_tracking]
        else:
            trackings = []
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    projects = LAC_ADF_Project.query.all()
    return render_template('admin/mla_sdf/approvals.html', base_template=base_template, trackings=trackings, projects=projects)

@app.route('/mla-sdf/execution')
@login_required
def mla_sdf_execution():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    executions = ProjectExecution.query.all()
    if not executions:
        project = LAC_ADF_Project.query.first()
        if project:
            default_execution = ProjectExecution(
                project_id=project.id,
                contractor_agency="Anil Constructions",
                work_order_amount=2450000.0,
                start_date="2026-06-20",
                scheduled_completion_date="2026-12-20",
                actual_completion_date="",
                physical_progress_pct=45.0,
                financial_progress_pct=40.0,
                amount_paid=1000000.0,
                running_bill_no="R-Bill-1",
                last_payment_date="2026-08-01",
                current_status="On Schedule",
                delay_reason="",
                revised_completion_date=""
            )
            db.session.add(default_execution)
            db.session.commit()
            executions = [default_execution]
        else:
            executions = []
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    projects = LAC_ADF_Project.query.all()
    return render_template('admin/mla_sdf/execution.html', base_template=base_template, executions=executions, projects=projects)

@app.route('/mla-sdf/payments')
@login_required
def mla_sdf_payments():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    payments = PaymentUtilization.query.all()
    if not payments:
        project = LAC_ADF_Project.query.first()
        if project:
            default_payment = PaymentUtilization(
                project_id=project.id,
                bill_no="BILL-8876",
                bill_date="2026-07-28",
                bill_amount=1000000.0,
                amount_approved=1000000.0,
                amount_paid=1000000.0,
                payment_date="2026-08-01",
                payment_reference="TXN998877",
                cumulative_expenditure=1000000.0,
                remaining_project_fund=1450000.0,
                uc_submitted=True,
                uc_date="2026-08-05"
            )
            db.session.add(default_payment)
            db.session.commit()
            payments = [default_payment]
        else:
            payments = []
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    projects = LAC_ADF_Project.query.all()
    return render_template('admin/mla_sdf/payments.html', base_template=base_template, payments=payments, projects=projects)

@app.route('/mla-sdf/monitoring')
@login_required
def mla_sdf_monitoring():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    inspections = MonitoringInspection.query.all()
    if not inspections:
        project = LAC_ADF_Project.query.first()
        if project:
            default_inspection = MonitoringInspection(
                inspection_id="INS-001",
                project_id=project.id,
                inspection_date="2026-07-15",
                inspection_officer="Officer Mathew",
                physical_progress_pct=30.0,
                financial_progress_pct=0.0,
                quality_status="Satisfactory",
                issues_identified="Minor delay in procurement",
                corrective_action="Accelerate sourcing",
                next_inspection_date="2026-08-20",
                inspection_report="",
                photographs="",
                remarks="Overall quality is good."
            )
            db.session.add(default_inspection)
            db.session.commit()
            inspections = [default_inspection]
        else:
            inspections = []
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    projects = LAC_ADF_Project.query.all()
    return render_template('admin/mla_sdf/monitoring.html', base_template=base_template, inspections=inspections, projects=projects)

@app.route('/mla-sdf/documents')
@login_required
def mla_sdf_documents():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    documents = MLA_SDF_Document.query.all()
    projects = LAC_ADF_Project.query.all()
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    return render_template('admin/mla_sdf/documents.html', base_template=base_template, documents=documents, projects=projects)

@app.route('/mla-sdf/reports')
@login_required
def mla_sdf_reports():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    reports = MLA_SDF_Report.query.all()
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    return render_template('admin/mla_sdf/reports.html', base_template=base_template, reports=reports)

@app.route('/mla-sdf/completion')
@login_required
def mla_sdf_completion():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    registers = CompletionAssetRegister.query.all()
    if not registers:
        project = LAC_ADF_Project.query.first()
        if project:
            default_register = CompletionAssetRegister(
                project_id=project.id,
                completion_certificate_no="CC-2026-004",
                completion_date="2026-12-24",
                final_cost=2450000.0,
                final_expenditure=2450000.0,
                asset_created="Community Hall",
                asset_location="East Fort",
                asset_custodian_department="Municipal Corporation",
                handover_date="2026-12-30",
                handover_document="",
                maintenance_responsibility="Municipal Corporation",
                maintenance_period="3 Years",
                asset_status="Functional"
            )
            db.session.add(default_register)
            db.session.commit()
            registers = [default_register]
        else:
            registers = []
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    projects = LAC_ADF_Project.query.all()
    return render_template('admin/mla_sdf/completion.html', base_template=base_template, registers=registers, projects=projects)

@app.route('/mla-sdf/asset-register')
@login_required
def mla_sdf_asset_register():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    registers = CompletionAssetRegister.query.all()
    projects = LAC_ADF_Project.query.all()
    base_template = 'dash_base_admin.html' if current_user.role == 'superadmin' else 'dash_base.html'
    return render_template('admin/mla_sdf/asset_register.html', base_template=base_template, registers=registers, projects=projects)
@app.route('/mla-sdf/fund-management/download')
@login_required
def download_fund_management():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    funds = FundManagement.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Fund Management"
    headers = ["SR", "Financial Year", "Constituency", "MLA Name", "Fund Type", "Allocation", "Previous Year Balance", "Total Available", "Amount Recommended", "Amount Sanctioned", "Amount Released", "Amount Utilized", "Balance Amount", "Total Projects", "Pending Projects", "Completed Projects"]
    ws.append(headers)
    for idx, f in enumerate(funds, 1):
        ws.append([idx, f.financial_year, f.constituency, f.mla_name, f.fund_type, f.annual_fund_allocation, f.previous_year_balance, f.total_available_fund, f.amount_recommended, f.amount_sanctioned, f.amount_released, f.amount_utilized, f.balance_amount, f.number_of_projects, f.pending_projects, f.completed_projects])
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"fund_management_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = Response(file_stream.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@app.route('/mla-sdf/proposals/download')
@login_required
def download_proposals():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    projects = LAC_ADF_Project.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Proposals"
    headers = ["SR", "Project ID", "Project Name", "Category", "Description", "Constituency", "Local Body", "Ward", "Location", "Beneficiary Institution", "Beneficiary Details", "MLA Recommendation Date", "Estimated Cost", "MLA ADF Amount", "Other Fund", "Implementing Department", "Implementing Agency", "Project Officer", "Priority", "Status"]
    ws.append(headers)
    for idx, p in enumerate(projects, 1):
        ws.append([idx, p.id, p.project_name, p.project_category, p.project_description, p.constituency, p.local_body, p.ward, p.location, p.beneficiary_institution, p.beneficiary_details, p.mla_recommendation_date, p.estimated_project_cost, p.mla_adf_amount, p.other_fund_contribution, p.implementing_department, p.implementing_agency, p.project_officer, p.priority, p.project_status])
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"proposals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = Response(file_stream.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@app.route('/mla-sdf/approvals/download')
@login_required
def download_approvals():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    trackings = ApprovalSanctionTracking.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Approvals"
    headers = ["SR", "Project ID", "Administrative Department", "MLA Recommendation", "Administrative Sanction No", "Administrative Sanction Date", "Technical Sanction No", "Technical Sanction Date", "Financial Concurrence No", "Financial Concurrence Date", "Detailed Estimate Prepared", "Tender Quotation Required", "Tender Date", "Work Order No", "Work Order Date", "Agreement No", "Agreement Date"]
    ws.append(headers)
    for idx, t in enumerate(trackings, 1):
        ws.append([idx, t.project_id, t.administrative_department, t.mla_recommendation, t.administrative_sanction_no, t.administrative_sanction_date, t.technical_sanction_no, t.technical_sanction_date, t.financial_concurrence_no, t.financial_concurrence_date, "Yes" if t.detailed_estimate_prepared else "No", "Yes" if t.tender_quotation_required else "No", t.tender_date, t.work_order_no, t.work_order_date, t.agreement_no, t.agreement_date])
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"approvals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = Response(file_stream.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@app.route('/mla-sdf/execution/download')
@login_required
def download_execution():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    executions = ProjectExecution.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Execution"
    headers = ["SR", "Project ID", "Contractor/Agency", "Work Order Amount", "Start Date", "Scheduled Completion Date", "Actual Completion Date", "Physical Progress %", "Financial Progress %", "Amount Paid", "Running Bill No", "Last Payment Date", "Current Status", "Delay Reason", "Revised Completion Date"]
    ws.append(headers)
    for idx, e in enumerate(executions, 1):
        ws.append([idx, e.project_id, e.contractor_agency, e.work_order_amount, e.start_date, e.scheduled_completion_date, e.actual_completion_date, e.physical_progress_pct, e.financial_progress_pct, e.amount_paid, e.running_bill_no, e.last_payment_date, e.current_status, e.delay_reason, e.revised_completion_date])
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"execution_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = Response(file_stream.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@app.route('/mla-sdf/payments/download')
@login_required
def download_payments():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    payments = PaymentUtilization.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    headers = ["SR", "Project ID", "Bill No", "Bill Date", "Bill Amount", "Amount Approved", "Amount Paid", "Payment Date", "Payment Reference", "Cumulative Expenditure", "Remaining Project Fund", "UC Submitted", "UC Date"]
    ws.append(headers)
    for idx, p in enumerate(payments, 1):
        ws.append([idx, p.project_id, p.bill_no, p.bill_date, p.bill_amount, p.amount_approved, p.amount_paid, p.payment_date, p.payment_reference, p.cumulative_expenditure, p.remaining_project_fund, "Yes" if p.uc_submitted else "No", p.uc_date])
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = Response(file_stream.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@app.route('/mla-sdf/monitoring/download')
@login_required
def download_monitoring():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    inspections = MonitoringInspection.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoring"
    headers = ["SR", "Inspection ID", "Project ID", "Inspection Date", "Inspection Officer", "Physical Progress %", "Financial Progress %", "Quality Status", "Issues Identified", "Corrective Action", "Next Inspection Date", "Inspection Report", "Photographs", "Remarks"]
    ws.append(headers)
    for idx, i in enumerate(inspections, 1):
        ws.append([idx, i.inspection_id, i.project_id, i.inspection_date, i.inspection_officer, i.physical_progress_pct, i.financial_progress_pct, i.quality_status, i.issues_identified, i.corrective_action, i.next_inspection_date, i.inspection_report, i.photographs, i.remarks])
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"monitoring_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = Response(file_stream.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@app.route('/mla-sdf/completion/download')
@login_required
def download_completion():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    registers = CompletionAssetRegister.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Completion"
    headers = ["SR", "Project ID", "Completion Certificate No", "Completion Date", "Final Cost", "Final Expenditure", "Asset Created", "Asset Location", "Asset Custodian Department", "Handover Date", "Handover Document", "Maintenance Responsibility", "Maintenance Period", "Asset Status"]
    ws.append(headers)
    for idx, r in enumerate(registers, 1):
        ws.append([idx, r.project_id, r.completion_certificate_no, r.completion_date, r.final_cost, r.final_expenditure, r.asset_created, r.asset_location, r.asset_custodian_department, r.handover_date, r.handover_document, r.maintenance_responsibility, r.maintenance_period, r.asset_status])
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"completion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = Response(file_stream.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route('/mla-sdf/proposals/add', methods=['POST'])
@login_required
def add_proposal():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    try:
        attachment_url = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                try:
                    upload_result = cloudinary.uploader.upload(file)
                    attachment_url = upload_result.get('secure_url')
                except Exception as e:
                    flash(f'File upload failed: {str(e)}', 'danger')

        project = LAC_ADF_Project(
            project_name=request.form.get('project_name'),
            project_category=request.form.get('project_category'),
            project_description=request.form.get('project_description'),
            constituency=request.form.get('constituency'),
            local_body=request.form.get('local_body'),
            ward=request.form.get('ward'),
            location=request.form.get('location'),
            beneficiary_institution=request.form.get('beneficiary_institution'),
            beneficiary_details=request.form.get('beneficiary_details'),
            mla_recommendation_date=request.form.get('mla_recommendation_date'),
            estimated_project_cost=float(request.form.get('estimated_project_cost', 0.0)),
            mla_adf_amount=float(request.form.get('mla_adf_amount', 0.0)),
            other_fund_contribution=float(request.form.get('other_fund_contribution', 0.0)),
            implementing_department=request.form.get('implementing_department'),
            implementing_agency=request.form.get('implementing_agency'),
            project_officer=request.form.get('project_officer'),
            priority=request.form.get('priority'),
            project_status=request.form.get('project_status', 'Proposed'),
            attachment_url=attachment_url
        )
        db.session.add(project)
        db.session.commit()
        flash(_('Project proposal added successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding proposal: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_proposals'))


@app.route('/mla-sdf/proposals/delete_attachment/<int:project_id>', methods=['POST'])
@login_required
def delete_proposal_attachment(project_id):
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    project = LAC_ADF_Project.query.get_or_404(project_id)
    try:
        project.attachment_url = None
        db.session.commit()
        flash(_('Attachment deleted successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting attachment: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_proposals'))

@app.route('/mla-sdf/approvals/add', methods=['POST'])
@login_required
def add_approval():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    try:
        attachment_url = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                try:
                    upload_result = cloudinary.uploader.upload(file)
                    attachment_url = upload_result.get('secure_url')
                except Exception as e:
                    flash(f'File upload failed: {str(e)}', 'danger')

        tracking = ApprovalSanctionTracking(
            project_id=int(request.form.get('project_id')),
            administrative_department=request.form.get('administrative_department'),
            mla_recommendation=request.form.get('mla_recommendation'),
            administrative_sanction_no=request.form.get('administrative_sanction_no'),
            administrative_sanction_date=request.form.get('administrative_sanction_date'),
            technical_sanction_no=request.form.get('technical_sanction_no'),
            technical_sanction_date=request.form.get('technical_sanction_date'),
            financial_concurrence_no=request.form.get('financial_concurrence_no'),
            financial_concurrence_date=request.form.get('financial_concurrence_date'),
            detailed_estimate_prepared=request.form.get('detailed_estimate_prepared') == 'true',
            tender_quotation_required=request.form.get('tender_quotation_required') == 'true',
            tender_date=request.form.get('tender_date'),
            work_order_no=request.form.get('work_order_no'),
            work_order_date=request.form.get('work_order_date'),
            agreement_no=request.form.get('agreement_no'),
            agreement_date=request.form.get('agreement_date'),
            attachment_url=attachment_url
        )
        db.session.add(tracking)
        db.session.commit()
        flash(_('Approval tracking record added successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding approval tracking: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_approvals'))


@app.route('/mla-sdf/approvals/delete_attachment/<int:rec_id>', methods=['POST'])
@login_required
def delete_approval_attachment(rec_id):
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    rec = ApprovalSanctionTracking.query.get_or_404(rec_id)
    try:
        rec.attachment_url = None
        db.session.commit()
        flash(_('Attachment deleted successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting attachment: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_approvals'))

@app.route('/mla-sdf/execution/add', methods=['POST'])
@login_required
def add_execution():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    try:
        attachment_url = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                try:
                    upload_result = cloudinary.uploader.upload(file)
                    attachment_url = upload_result.get('secure_url')
                except Exception as e:
                    flash(f'File upload failed: {str(e)}', 'danger')

        execution = ProjectExecution(
            project_id=int(request.form.get('project_id')),
            contractor_agency=request.form.get('contractor_agency'),
            work_order_amount=float(request.form.get('work_order_amount', 0.0)),
            start_date=request.form.get('start_date'),
            scheduled_completion_date=request.form.get('scheduled_completion_date'),
            actual_completion_date=request.form.get('actual_completion_date'),
            physical_progress_pct=float(request.form.get('physical_progress_pct', 0.0)),
            financial_progress_pct=float(request.form.get('financial_progress_pct', 0.0)),
            amount_paid=float(request.form.get('amount_paid', 0.0)),
            running_bill_no=request.form.get('running_bill_no'),
            last_payment_date=request.form.get('last_payment_date'),
            current_status=request.form.get('current_status'),
            delay_reason=request.form.get('delay_reason'),
            revised_completion_date=request.form.get('revised_completion_date'),
            attachment_url=attachment_url
        )
        db.session.add(execution)
        db.session.commit()
        flash(_('Project execution record added successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding execution record: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_execution'))


@app.route('/mla-sdf/execution/delete_attachment/<int:rec_id>', methods=['POST'])
@login_required
def delete_execution_attachment(rec_id):
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    rec = ProjectExecution.query.get_or_404(rec_id)
    try:
        rec.attachment_url = None
        db.session.commit()
        flash(_('Attachment deleted successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting attachment: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_execution'))

@app.route('/mla-sdf/payments/add', methods=['POST'])
@login_required
def add_payment():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    try:
        attachment_url = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                try:
                    upload_result = cloudinary.uploader.upload(file)
                    attachment_url = upload_result.get('secure_url')
                except Exception as e:
                    flash(f'File upload failed: {str(e)}', 'danger')

        payment = PaymentUtilization(
            project_id=int(request.form.get('project_id')),
            bill_no=request.form.get('bill_no'),
            bill_date=request.form.get('bill_date'),
            bill_amount=float(request.form.get('bill_amount', 0.0)),
            amount_approved=float(request.form.get('amount_approved', 0.0)),
            amount_paid=float(request.form.get('amount_paid', 0.0)),
            payment_date=request.form.get('payment_date'),
            payment_reference=request.form.get('payment_reference'),
            cumulative_expenditure=float(request.form.get('cumulative_expenditure', 0.0)),
            remaining_project_fund=float(request.form.get('remaining_project_fund', 0.0)),
            uc_submitted=request.form.get('uc_submitted') == 'true',
            uc_date=request.form.get('uc_date'),
            attachment_url=attachment_url
        )
        db.session.add(payment)
        db.session.commit()
        flash(_('Payment record added successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding payment: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_payments'))


@app.route('/mla-sdf/payments/delete_attachment/<int:rec_id>', methods=['POST'])
@login_required
def delete_payment_attachment(rec_id):
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    rec = PaymentUtilization.query.get_or_404(rec_id)
    try:
        rec.attachment_url = None
        db.session.commit()
        flash(_('Attachment deleted successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting attachment: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_payments'))

@app.route('/mla-sdf/monitoring/add', methods=['POST'])
@login_required
def add_monitoring():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    try:
        attachment_url = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                try:
                    upload_result = cloudinary.uploader.upload(file)
                    attachment_url = upload_result.get('secure_url')
                except Exception as e:
                    flash(f'File upload failed: {str(e)}', 'danger')

        inspection = MonitoringInspection(
            inspection_id=request.form.get('inspection_id'),
            project_id=int(request.form.get('project_id')),
            inspection_date=request.form.get('inspection_date'),
            inspection_officer=request.form.get('inspection_officer'),
            physical_progress_pct=float(request.form.get('physical_progress_pct', 0.0)),
            financial_progress_pct=float(request.form.get('financial_progress_pct', 0.0)),
            quality_status=request.form.get('quality_status'),
            issues_identified=request.form.get('issues_identified'),
            corrective_action=request.form.get('corrective_action'),
            next_inspection_date=request.form.get('next_inspection_date'),
            remarks=request.form.get('remarks'),
            attachment_url=attachment_url
        )
        db.session.add(inspection)
        db.session.commit()
        flash(_('Inspection log added successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding inspection log: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_monitoring'))


@app.route('/mla-sdf/monitoring/delete_attachment/<int:rec_id>', methods=['POST'])
@login_required
def delete_monitoring_attachment(rec_id):
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    rec = MonitoringInspection.query.get_or_404(rec_id)
    try:
        rec.attachment_url = None
        db.session.commit()
        flash(_('Attachment deleted successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting attachment: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_monitoring'))

@app.route('/mla-sdf/completion/add', methods=['POST'])
@login_required
def add_completion():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    try:
        attachment_url = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                try:
                    upload_result = cloudinary.uploader.upload(file)
                    attachment_url = upload_result.get('secure_url')
                except Exception as e:
                    flash(f'File upload failed: {str(e)}', 'danger')

        register = CompletionAssetRegister(
            project_id=int(request.form.get('project_id')),
            completion_certificate_no=request.form.get('completion_certificate_no'),
            completion_date=request.form.get('completion_date'),
            final_cost=float(request.form.get('final_cost', 0.0)),
            final_expenditure=float(request.form.get('final_expenditure', 0.0)),
            asset_created=request.form.get('asset_created'),
            asset_location=request.form.get('asset_location'),
            asset_custodian_department=request.form.get('asset_custodian_department'),
            handover_date=request.form.get('handover_date'),
            maintenance_responsibility=request.form.get('maintenance_responsibility'),
            maintenance_period=request.form.get('maintenance_period'),
            asset_status=request.form.get('asset_status'),
            attachment_url=attachment_url
        )
        db.session.add(register)
        db.session.commit()
        flash(_('Completion and asset register record added successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding completion record: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_completion'))


@app.route('/mla-sdf/completion/delete_attachment/<int:rec_id>', methods=['POST'])
@login_required
def delete_completion_attachment(rec_id):
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    rec = CompletionAssetRegister.query.get_or_404(rec_id)
    try:
        rec.attachment_url = None
        db.session.commit()
        flash(_('Attachment deleted successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting attachment: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_completion'))


@app.route('/mla-sdf/documents/add', methods=['POST'])
@login_required
def add_document():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    try:
        attachment_url = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                try:
                    upload_result = cloudinary.uploader.upload(file)
                    attachment_url = upload_result.get('secure_url')
                except Exception as e:
                    flash(f'File upload failed: {str(e)}', 'danger')

        doc = MLA_SDF_Document(
            project_id=int(request.form.get('project_id')),
            document_title=request.form.get('document_title'),
            document_category=request.form.get('document_category'),
            uploaded_date=request.form.get('uploaded_date'),
            remarks=request.form.get('remarks'),
            attachment_url=attachment_url
        )
        db.session.add(doc)
        db.session.commit()
        flash(_('Document uploaded successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding document: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_documents'))


@app.route('/mla-sdf/documents/delete_attachment/<int:rec_id>', methods=['POST'])
@login_required
def delete_document_attachment(rec_id):
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    rec = MLA_SDF_Document.query.get_or_404(rec_id)
    try:
        rec.attachment_url = None
        db.session.commit()
        flash(_('Attachment deleted successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting attachment: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_documents'))


@app.route('/mla-sdf/documents/download')
@login_required
def download_documents():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    documents = MLA_SDF_Document.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"
    headers = ["ID", "Project ID", "Document Title", "Document Category", "Uploaded Date", "Remarks", "Attachment URL"]
    ws.append(headers)
    for d in documents:
        ws.append([d.id, d.project_id, d.document_title, d.document_category, d.uploaded_date, d.remarks, d.attachment_url])
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"MLA_SDF_Documents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = Response(file_stream.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route('/mla-sdf/reports/add', methods=['POST'])
@login_required
def add_report():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    try:
        attachment_url = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                try:
                    upload_result = cloudinary.uploader.upload(file)
                    attachment_url = upload_result.get('secure_url')
                except Exception as e:
                    flash(f'File upload failed: {str(e)}', 'danger')

        rep = MLA_SDF_Report(
            report_name=request.form.get('report_name'),
            report_type=request.form.get('report_type'),
            generated_by=request.form.get('generated_by'),
            financial_year=request.form.get('financial_year'),
            generated_date=request.form.get('generated_date'),
            description=request.form.get('description'),
            attachment_url=attachment_url
        )
        db.session.add(rep)
        db.session.commit()
        flash(_('Report record added successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding report: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_reports'))


@app.route('/mla-sdf/reports/delete_attachment/<int:rec_id>', methods=['POST'])
@login_required
def delete_report_attachment(rec_id):
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    rec = MLA_SDF_Report.query.get_or_404(rec_id)
    try:
        rec.attachment_url = None
        db.session.commit()
        flash(_('Attachment deleted successfully.'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting attachment: {str(e)}', 'danger')
    return redirect(url_for('mla_sdf_reports'))


@app.route('/mla-sdf/reports/download')
@login_required
def download_reports():
    if current_user.role not in ['admin', 'superadmin']:
        return redirect(url_for('index'))
    reports = MLA_SDF_Report.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"
    headers = ["ID", "Report Name", "Report Type", "Generated By", "Financial Year", "Generated Date", "Description", "Attachment URL"]
    ws.append(headers)
    for r in reports:
        ws.append([r.id, r.report_name, r.report_type, r.generated_by, r.financial_year, r.generated_date, r.description, r.attachment_url])
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"MLA_SDF_Reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = Response(file_stream.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


# --- Setup Script ---
@app.cli.command("init-db")
def init_db():
    db.create_all()
    # Create super admin
    if not User.query.filter_by(username='superadmin').first():
        sa = User(username='superadmin', role='superadmin')
        sa.set_password('admin123')
        db.session.add(sa)
    # Create some departments
    if not Department.query.first():
        for d in ['IT', 'HR', 'Finance', 'Operations']:
            db.session.add(Department(name=d))
    db.session.commit()
    print("Database initialized.")

if __name__ == '__main__':
    print("DB URI:", app.config['SQLALCHEMY_DATABASE_URI'])
    app.run(debug=True)
